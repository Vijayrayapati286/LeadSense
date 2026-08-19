"""Read LinkedIn URLs from uploaded spreadsheets and write enriched results."""

from __future__ import annotations

import io
import logging
import re
from pathlib import Path
from typing import Any

import pandas as pd

from app.config import get_settings
from app.linkedin.bulk_models import ITEM_SUCCESS
from app.linkedin.validator import is_linkedin_in_profile_url, normalize_profile_url, validate_profile_url
from app.services.excel_service import SUPPORTED_EXTENSIONS, _normalize_header, _read_dataframe

logger = logging.getLogger(__name__)

LINKEDIN_IN_RE = re.compile(
    r"https?://(?:www\.)?linkedin\.com/in/[^/?#\s\"']+",
    re.I,
)

URL_COLUMN_ALIASES = {
    "linkedinurl",
    "linkedin",
    "profileurl",
    "url",
    "link",
    "linkedinprofileurl",
    "linkedinprofile",
    "linkedinprofilelink",
    "linkedinlink",
    "personlinkedinurl",
    "personlinkedin",
}

OUTPUT_COLUMNS = (
    "LinkedIn URL",
    "Name",
    "Headline",
    "Company",
    "Job Title",
    "Location",
    "Summary",
    "Followers",
    "Connections",
    "Status",
    "Error",
)

OUTPUTS_DIR = Path(__file__).resolve().parent / "outputs"


def _max_bulk_urls() -> int:
    return max(get_settings().max_bulk_urls, 1)


class BulkExcelError(ValueError):
    """User-facing bulk Excel parse/validation error."""


class BulkExcelService:
    def read_upload(self, file_content: bytes, filename: str) -> pd.DataFrame:
        suffix = Path(filename or "").suffix.lower()
        if suffix not in SUPPORTED_EXTENSIONS:
            supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
            raise BulkExcelError(
                f"Unsupported file type '{suffix or 'unknown'}'. Supported: {supported}"
            )
        try:
            return _read_dataframe(file_content, filename)
        except ValueError as exc:
            raise BulkExcelError(str(exc)) from exc

    def extract_url_rows(self, df: pd.DataFrame) -> list[dict[str, Any]]:
        """Return one dict per spreadsheet row that has a URL (including duplicates)."""
        url_col = self._find_url_column(df)
        rows: list[dict[str, Any]] = []
        seen: dict[str, int] = {}
        columns = [str(c) for c in df.columns]

        for pos, (idx, series) in enumerate(df.iterrows()):
            raw = self._url_from_row(series, url_col)
            extra = {}
            for col in columns:
                extra[col] = self._cell(series[col] if col in series.index else None)
            if not raw:
                rows.append(
                    {
                        "row_index": int(idx) if str(idx).isdigit() or isinstance(idx, int) else pos,
                        "source_row_number": pos + 1,
                        "raw_url": "",
                        "normalized_url": "",
                        "is_valid": False,
                        "is_duplicate": False,
                        "error": "No LinkedIn profile URL on this row",
                        "source_row_json": extra,
                        "input_columns": columns,
                    }
                )
                continue
            error = None
            is_valid = True
            try:
                normalized = validate_profile_url(raw)
            except ValueError as exc:
                is_valid = False
                normalized = normalize_profile_url(raw) if is_linkedin_in_profile_url(raw) else ""
                error = str(exc)
            duplicate_of_row = seen.get(normalized) if normalized else None
            if normalized and normalized not in seen:
                seen[normalized] = pos
            rows.append(
                {
                    "row_index": int(idx) if str(idx).isdigit() or isinstance(idx, int) else pos,
                    "source_row_number": pos + 1,
                    "raw_url": raw,
                    "normalized_url": normalized,
                    "is_valid": is_valid,
                    "is_duplicate": duplicate_of_row is not None,
                    "error": error,
                    "source_row_json": extra,
                    "input_columns": columns,
                }
            )

        if not any(r.get("normalized_url") or r.get("raw_url") for r in rows):
            raise BulkExcelError(
                "No LinkedIn profile URLs found. Add a column like 'LinkedIn URL' "
                "with links such as https://www.linkedin.com/in/username/"
            )
        max_urls = _max_bulk_urls()
        if len(rows) > max_urls:
            raise BulkExcelError(
                f"Too many URLs ({len(rows)}). Maximum allowed per upload is {max_urls}."
            )
        return rows

    def input_columns(self, df: pd.DataFrame) -> list[str]:
        return [str(c) for c in df.columns]

    def build_result_workbook(
        self,
        *,
        job_id: str,
        url_rows: list[dict[str, Any]],
        results_by_url: dict[str, dict[str, Any]],
        processed_only: bool = False,
    ) -> tuple[bytes, str, str]:
        """Write profile rows to outputs/bulk_{job_id}.xlsx (overwritten on each save)."""
        rows_to_write = url_rows
        if processed_only:
            rows_to_write = [
                entry
                for entry in url_rows
                if entry["normalized_url"] in results_by_url
            ]

        out_rows: list[dict[str, Any]] = []
        for entry in rows_to_write:
            url = entry["normalized_url"]
            result = results_by_url.get(url, {})
            data = result.get("data") or {}
            status = result.get("status", "pending" if not processed_only else "failed")
            error = result.get("error") or ""

            out_rows.append(
                {
                    "LinkedIn URL": url,
                    "Name": self._cell(data.get("name")),
                    "Headline": self._cell(data.get("headline")),
                    "Company": self._cell(data.get("company")),
                    "Job Title": self._cell(data.get("job_title")),
                    "Location": self._cell(data.get("location")),
                    "Summary": self._cell(data.get("summary")),
                    "Followers": data.get("followers") if data.get("followers") is not None else "",
                    "Connections": data.get("connections") if data.get("connections") is not None else "",
                    "Status": status,
                    "Error": error or None,
                }
            )

        result_df = pd.DataFrame(out_rows, columns=list(OUTPUT_COLUMNS))
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            result_df.to_excel(writer, index=False, sheet_name="Profiles")

            # Style rows: mark any profile that wasn't extracted (`status != "ok"`)
            # as red so users can immediately see which URLs still need retry.
            from openpyxl.styles import Font, PatternFill

            ws = writer.sheets.get("Profiles")
            if ws is not None:
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                red_font = Font(color="9C0006")

                # Excel rows start at 1; header is at row=1, first data row is row=2.
                for i, out in enumerate(out_rows):
                    excel_row = i + 2
                    status = str(out.get("Status") or "").strip().lower()
                    if status != "ok":
                        # Color the full row (all columns) red.
                        for excel_col in range(1, len(OUTPUT_COLUMNS) + 1):
                            cell = ws.cell(row=excel_row, column=excel_col)
                            cell.fill = red_fill
                            cell.font = red_font

        content = buffer.getvalue()
        filename = self.job_filename(job_id)
        relative = f"outputs/{filename}"

        try:
            OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
            (OUTPUTS_DIR / filename).write_bytes(content)
            logger.info(
                "Saved bulk LinkedIn Excel to outputs/%s (%s rows, partial=%s)",
                filename,
                len(out_rows),
                processed_only,
            )
        except OSError:
            logger.exception("Could not save bulk LinkedIn Excel to disk")

        return content, filename, relative

    def build_result_workbook_from_items(
        self,
        *,
        job_id: str,
        items: list[Any],
        input_columns: list[str],
    ) -> tuple[bytes, str, str]:
        from app.linkedin.verification import original_fields

        columns = [
            "Original Name",
            "Original LinkedIn URL",
            "Original Designation",
            "Original Company",
            "Original Location",
            "Extracted Name",
            "Extracted Designation",
            "Extracted Company",
            "Extracted Location",
            "Extracted About",
            "Extraction Status",
            "Extraction Attempts",
            "Name Match",
            "Designation Match",
            "Company Match",
            "Location Match",
            "Verification Score",
            "Verification Status",
            "Verification Reason",
            "Error",
        ]
        extra_original = [
            c
            for c in (input_columns or [])
            if c
            and _normalize_header(c)
            not in {
                "name",
                "fullname",
                "linkedinurl",
                "linkedin",
                "profileurl",
                "url",
                "link",
                "linkedinprofileurl",
                "designation",
                "jobtitle",
                "title",
                "company",
                "location",
            }
        ]
        columns = extra_original + columns

        def _match_label(value: Any) -> str:
            if value is True:
                return "true"
            if value is False:
                return "false"
            return ""

        out_rows: list[dict[str, Any]] = []
        for item in items:
            source = item.source_row_json if isinstance(getattr(item, "source_row_json", None), dict) else {}
            originals = original_fields(source)
            row: dict[str, Any] = {}
            for col in extra_original:
                row[col] = source.get(col)
            row["Original Name"] = originals.get("name")
            row["Original LinkedIn URL"] = getattr(item, "normalized_url", None) or getattr(
                item, "profile_url", None
            )
            row["Original Designation"] = originals.get("designation")
            row["Original Company"] = originals.get("company")
            row["Original Location"] = originals.get("location")
            row["Extracted Name"] = getattr(item, "name", None)
            row["Extracted Designation"] = getattr(item, "designation", None)
            row["Extracted Company"] = getattr(item, "company", None)
            row["Extracted Location"] = getattr(item, "location", None)
            row["Extracted About"] = getattr(item, "about", None)
            row["Extraction Status"] = getattr(item, "status", "") or ""
            row["Extraction Attempts"] = getattr(item, "attempt_count", 0) or 0
            row["Name Match"] = _match_label(getattr(item, "name_match", None))
            row["Designation Match"] = _match_label(getattr(item, "designation_match", None))
            row["Company Match"] = _match_label(getattr(item, "company_match", None))
            row["Location Match"] = _match_label(getattr(item, "location_match", None))
            score = getattr(item, "verification_score", None)
            row["Verification Score"] = "" if score is None else score
            row["Verification Status"] = getattr(item, "verification_status", "") or ""
            row["Verification Reason"] = getattr(item, "verification_reason", None)
            row["Error"] = getattr(item, "last_error", None)
            out_rows.append(row)

        result_df = pd.DataFrame(out_rows, columns=columns)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            result_df.to_excel(writer, index=False, sheet_name="Profiles")
            from openpyxl.styles import Font, PatternFill

            ws = writer.sheets.get("Profiles")
            if ws is not None:
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                red_font = Font(color="9C0006")
                for i, out in enumerate(out_rows):
                    excel_row = i + 2
                    status = str(out.get("Extraction Status") or out.get("Status") or "").strip().upper()
                    if status and status != ITEM_SUCCESS and status != "OK":
                        for excel_col in range(1, len(columns) + 1):
                            cell = ws.cell(row=excel_row, column=excel_col)
                            cell.fill = red_fill
                            cell.font = red_font

        content = buffer.getvalue()
        filename = self.job_filename(job_id)
        relative = f"outputs/{filename}"
        try:
            OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
            (OUTPUTS_DIR / filename).write_bytes(content)
        except OSError:
            logger.exception("Could not save bulk LinkedIn Excel to disk")
        return content, filename, relative

    @staticmethod
    def job_filename(job_id: str) -> str:
        return f"bulk_{job_id}.xlsx"

    def resolve_safe_path(self, filename: str) -> Path | None:
        name = (filename or "").strip()
        if not name or "/" in name or "\\" in name or ".." in name:
            return None
        if not re.fullmatch(r"bulk_[0-9a-f-]{36}\.xlsx", name) and not re.fullmatch(
            r"bulk_profile_\d{8}_\d{6}\.xlsx", name
        ):
            return None
        path = (OUTPUTS_DIR / name).resolve()
        try:
            path.relative_to(OUTPUTS_DIR.resolve())
        except ValueError:
            return None
        if not path.is_file():
            return None
        return path

    def _find_url_column(self, df: pd.DataFrame) -> str | None:
        for col in df.columns:
            if _normalize_header(str(col)) in URL_COLUMN_ALIASES:
                return str(col)

        best_col: str | None = None
        best_count = 0
        for col in df.columns:
            count = sum(1 for v in df[col] if self._extract_url_from_text(v))
            if count > best_count:
                best_count = count
                best_col = str(col)
        return best_col if best_count > 0 else None

    def _url_from_row(self, series: pd.Series, url_col: str | None) -> str:
        if url_col and url_col in series.index:
            found = self._extract_url_from_text(series[url_col])
            if found:
                return found
        for value in series.values:
            found = self._extract_url_from_text(value)
            if found:
                return found
        return ""

    @staticmethod
    def _extract_url_from_text(value: Any) -> str:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        text = str(value).strip()
        if not text:
            return ""
        if is_linkedin_in_profile_url(text):
            return text
        match = LINKEDIN_IN_RE.search(text)
        return match.group(0) if match else ""

    @staticmethod
    def _cell(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text or text.lower() in {"none", "null", "n/a"}:
            return None
        return text
